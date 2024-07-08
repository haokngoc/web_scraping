from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time

# Hàm đọc các ID từ Excel
def read_ids_from_excel(file_path, sheet_name, column_index):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    ids = []
    for row in sheet.iter_rows(values_only=True):
        ids.append(row[column_index])
    return ids

def search_on_mouser(id):
    options = webdriver.ChromeOptions()
    options.add_argument("window-size=1200x600")  
    
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.mouser.vn")
    
    try:

        time.sleep(5)
        

        search_box = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "as-input-024"))
        )
        search_box.send_keys(id)
        search_box.send_keys(Keys.RETURN)
        
        # Chờ kết quả tìm kiếm và in ra tiêu đề của kết quả đầu tiên
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.productDetail")))
        first_result = driver.find_element(By.CSS_SELECTOR, "div.productDetail")
        print(first_result.text)  # In ra nội dung của kết quả đầu tiên
        
    finally:
        driver.quit()

if __name__ == "__main__":
    excel_file = "C:/Users/Admin/Documents/Zalo Received Files/Bom19182.xlsx"  
    sheet_name = "Raw"  
    id_column_index = 0  

    ids = read_ids_from_excel(excel_file, sheet_name, id_column_index)
    
    for id in ids:
        print(f"Đang xử lý ID: {id}")
        search_on_mouser(id)
